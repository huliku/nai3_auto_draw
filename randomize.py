import openpyxl
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import time
import threading
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC

# 创建全局锁对象
lock = threading.Lock()

def run_timer():
    start_time = time.time()

    try:
        while True:
            # 计算当前运行时间
            current_time = time.time() - start_time
            minutes_elapsed = int(current_time // 60)
            seconds_elapsed = int(current_time % 60)
            # 打印已经运行的时间
            print(f"已经运行时间：{minutes_elapsed} 分钟 {seconds_elapsed} 秒")
            # 更新时间，单位为秒
            time.sleep(60)

    finally:
        # 记录程序结束时间
        end_time = time.time()
        total_time = end_time - start_time
        # 打印程序总运行时间
        print(f"程序总运行时间：{int(total_time // 60)} 分钟 {int(total_time % 60)} 秒")

def run_webdriver():
    # 配置Chrome浏览器
    options = webdriver.ChromeOptions()
    options.add_experimental_option('detach', True)
    options.add_argument('--start-maximized')
    driver = webdriver.Chrome(options=options)

    # 打开目标网站
    driver.get('https://novelai.net/image')

    # 等待进入主界面
    time.sleep(20)

    # 输入Email
    inputTag = driver.find_element(By.ID, 'username')
    inputTag.send_keys("xxxxxxxx")

    # 输入Password
    inputTag = driver.find_element(By.ID, 'password')
    inputTag.send_keys("xxxxxxxx")

    # 点击Sign In，直到进入下个页面
    old_url = driver.current_url
    while True:
        try:
            # 查找按钮并点击
            driver.find_element(By.XPATH, '//*[@id="__next"]/div/form/div[1]/input[3]').click()
            WebDriverWait(driver, 8).until(EC.url_changes(old_url))
            break
        except:
            pass
    #等待网页响应
    time.sleep(20)

    # 进入生图
    driver.find_element(By.XPATH, '//*[@id="app"]/div[3]/div[4]/div[2]/div/div[3]/button[1]').click()

    # 设置等待时间
    time.sleep(3)

    # 输入负面tag
    inputTag = driver.find_element(By.XPATH, '//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[2]/div/div[4]/textarea')
    inputTag.send_keys("lowres, {bad}, error, fewer, extra, missing, worst quality, jpeg artifacts, bad quality, watermark, unfinished, displeasing, chromatic aberration, signature, extra digits, artistic error, username, scan, [abstract],mosaic censoring,bar censor,")

    # 指定Excel文件路径
    excel_file_path = './Prompt/tag文本.xlsx'

    # 定义点击按钮的函数
    def click_button():
        count = 1
        clicks_per_generation = 10

        while True:
            try:
                # 尝试点击按钮
                driver.find_element(By.XPATH, '//*[@id="__next"]/div[2]/div[4]/div[1]/div[5]/button').click()
                # 输出点击次数
                if count % clicks_per_generation == 0:
                    print(f"生成点击次数: {count}")
                # 等待一段时间
                time.sleep(13)
                # 增加计数
                count += 1

            except Exception as e:
                # 不输出异常信息，只pass
                pass

    # 启动点击按钮的线程
    thread_click_button = threading.Thread(target=click_button)
    thread_click_button.start()

    # 启动计时器线程
    thread_timer = threading.Thread(target=run_timer)
    thread_timer.start()

    def 读取并选择随机行(sheet):
        随机行 = random.randint(1, sheet.max_row)
        数据A = sheet.cell(row=随机行, column=1).value
        数据B = sheet.cell(row=随机行, column=2).value
        return 数据A, 数据B

    def 读取工作表并选择随机数据(workbook, sheet_name):
        sheet = workbook[sheet_name]
        数据A, 数据B = 读取并选择随机行(sheet)
        return 数据A, 数据B

    try:

        while True:
            try:
                with lock:  # 使用锁确保两个线程不会同时执行关键部分的代码
                    # 定位Prompt,及分辨率框,加入循环是为了防止网页刷新无法定位。
                    inputTag = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[2]/div/div[2]/textarea')
                    inputTag_width = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[3]/div/div[2]/div[2]/input[1]')
                    inputTag_height = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[3]/div/div[2]/div[2]/input[2]')

                    # 打开Excel文件
                    workbook = openpyxl.load_workbook(excel_file_path)

                    # 写入你需要读取的Sheet，按照已有格式修改
                    角色A, 角色B = 读取工作表并选择随机数据(workbook, '角色')
                    画风A, 画风B = 读取工作表并选择随机数据(workbook, '画风')
                    服装A, 服装B = 读取工作表并选择随机数据(workbook, '服装')
                    表情A, 表情B = 读取工作表并选择随机数据(workbook, '表情')
                    动作A, 动作B = 读取工作表并选择随机数据(workbook, '动作')
                    # 场景A, 场景B = 读取工作表并选择随机数据(workbook, '场景')
                    单独场景A, 单独场景B = 读取工作表并选择随机数据(workbook, '单独场景')

                    # 此处用于读取随机分辨率
                    分辨率A, 分辨率B = 读取工作表并选择随机数据(workbook, '分辨率')

                    # TAG输入组成
                    # 在这调整你的Prompt输入顺序，请注意检查，不要把B复制成A
                    Prompt顺序 = [角色B, 画风B, 服装B, 表情B, 动作B, 单独场景B]
                    input_text = "".join(Prompt顺序)

                    # 以下是设置随机分辨率,不用动,先点击输入框,等待一秒,再填入分辨率。
                    inputTag_width.click()
                    time.sleep(1)
                    inputTag_width.send_keys(str(分辨率A))
                    inputTag_width.click()
                    time.sleep(1)
                    inputTag_height.send_keys(str(分辨率B))

                    # 关闭Excel文件
                    workbook.close()

                    # 点击输入框等待1秒
                    inputTag.click()
                    time.sleep(1)
                    # 将组合好的文本输入到文本框
                    inputTag.send_keys(Keys.CONTROL + "a")
                    inputTag.send_keys(input_text)

                    # 在这输出你的prompt注释，不想输出可以在print前加入# 号不输出，可以少写，例如不写画风，自行选择
                    print("————————————————————————————————————————————————————————————————")
                    print("本次随机结果是：")
                    print(f"{角色A}{服装A}{表情A}{动作A}{单独场景A}")
                    print(f"生成分辨率：{分辨率A} x {分辨率B}")
                    print("————————————————————————————————————————————————————————————————")

                    # 在这修改随机循环时间
                    time.sleep(60)

            except Exception as e:
                print(f"发生异常: {e}")
                # 继续下一次循环
                continue

    except Exception as e:
        print(f"发生异常: {e}")

    finally:
        # 等待点击按钮的线程结束
        thread_click_button.join()
        # 等待计时器线程结束
        thread_timer.join()

if __name__ == '__main__':
    run_webdriver()
