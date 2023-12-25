import openpyxl
import random
import os
import sys
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.alert import Alert
import time
import threading
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC

global description
description = "Randomized"
global input_text
input_text = "Randomized"
global resolution_A
global resolution_B
resolution_A = 0  # 初始值
resolution_B = 0  # 初始值

def run_timer():
    start_time = time.time()
    try:
        while True:
            # 计算当前运行时间
            current_time = time.time() - start_time
            minutes_elapsed = int(current_time // 60)
            seconds_elapsed = int(current_time % 60)
            # 打印已经运行的时间
            print(f"已经运行时间：{minutes_elapsed} 分钟 {seconds_elapsed} 秒")
            # 更新时间，单位为秒
            time.sleep(60)

    finally:
        # 记录程序结束时间
        end_time = time.time()
        total_time = end_time - start_time
        # 打印程序总运行时间
        print(f"程序总运行时间：{int(total_time // 60)} 分钟 {int(total_time % 60)} 秒")


def run_webdriver():
    # 配置Chrome浏览器
    options = webdriver.ChromeOptions()
    options.add_experimental_option('detach', True)
    options.add_argument('--start-maximized')
    # 允许修改默认下载路径，用于移动下载压缩包位置，使用绝对路径例如：E:\\nai3_auto_draw\\novelaitemp
    prefs = {"download.default_directory": 'E:\\nai3自动化\\novelaitemp'}
    if not os.path.exists('E:\\nai3自动化\\novelaitemp'):
        os.mkdir('E:\\nai3自动化\\novelaitemp')
    # add_experimental_option表示将这些首选项添加到他们的Selenium Webdriver对象中
    options.add_experimental_option("prefs", prefs)
    # 按照设置的首选项进行浏览器启动
    driver = webdriver.Chrome(options=options)

    # 打开目标网站
    driver.get('https://novelai.net/image')

    # 等待进入主界面
    time.sleep(20)

    # 输入Email
    inputTag = driver.find_element(By.ID, 'username')
    inputTag.send_keys("xxxxxxxxxxx")

    # 输入Password
    inputTag = driver.find_element(By.ID, 'password')
    inputTag.send_keys("xxxxxxxxxxx")

    # 点击Sign In，直到进入下个页面
    old_url = driver.current_url
    while True:
        try:
            # 查找按钮并点击
            driver.find_element(By.XPATH, '//*[@id="__next"]/div/form/div[1]/input[3]').click()
            WebDriverWait(driver, 8).until(EC.url_changes(old_url))
            break
        except:
            pass
    #等待网页响应
    time.sleep(20)

    # 进入生图
    driver.find_element(By.XPATH, '//*[@id="app"]/div[3]/div[4]/div[2]/div/div[3]/button[1]').click()

    # 设置等待时间
    time.sleep(3)

    # 输入负面tag
    inputTag = driver.find_element(By.XPATH, '//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[2]/div/div[4]/textarea')
    inputTag.send_keys("lowres, {bad}, error, fewer, extra, missing, worst quality, jpeg artifacts, bad quality, watermark, unfinished, displeasing, chromatic aberration, signature, extra digits, artistic error, username, scan, [abstract],mosaic censoring,bar censor,")

    # 指定Excel文件路径
    excel_file_path = './Prompt/tag文本.xlsx'

    #随机读取excel并输入到咒语输入框中的函数
    def read_text():
        def 读取并选择随机行(sheet):
            随机行 = random.randint(1, sheet.max_row)
            数据A = sheet.cell(row=随机行, column=1).value
            数据B = sheet.cell(row=随机行, column=2).value
            return 数据A, 数据B

        def 读取工作表并选择随机数据(workbook, sheet_name):
            sheet = workbook[sheet_name]
            数据A, 数据B = 读取并选择随机行(sheet)
            return 数据A, 数据B
        try:
            while True:
                try:

                    inputTag = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[2]/div/div[2]/textarea')
                    inputTag_width = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[3]/div/div[2]/div[2]/input[1]')
                    inputTag_height = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[3]/div/div[2]/div[2]/input[2]')
                    # 打开Excel文件
                    workbook = openpyxl.load_workbook(excel_file_path)

                    # 写入你需要读取的Sheet，按照已有格式修改
                    角色A, 角色B = 读取工作表并选择随机数据(workbook, '角色')
                    画风A, 画风B = 读取工作表并选择随机数据(workbook, '画风')
                    服装A, 服装B = 读取工作表并选择随机数据(workbook, '服装')
                    表情A, 表情B = 读取工作表并选择随机数据(workbook, '表情')
                    动作A, 动作B = 读取工作表并选择随机数据(workbook, '动作')
                    场景A, 场景B = 读取工作表并选择随机数据(workbook, '场景')
                    分辨率A, 分辨率B = 读取工作表并选择随机数据(workbook, '分辨率')
                    # TAG输入组成
                    # 在这调整你的Prompt输入顺序，请注意检查，不要把B复制成A
                    Prompt顺序 = [角色B,服装B, 画风B, 表情B, 动作B, 场景B]
                    global input_text
                    input_text = "".join(Prompt顺序)
                    global description
                    # 在这调整你的prompt注释及压缩包命名格式
                    description =f'{角色A}_{画风A}_{服装A}-{表情A}_{动作A}_{场景A}_{分辨率A}X{分辨率B}'

                    # 将分辨率A和分辨率B赋值给全局变量
                    global resolution_A
                    global resolution_B
                    resolution_A = 分辨率A
                    resolution_B = 分辨率B

                    # 以下是设置随机分辨率，不用动
                    inputTag_width.click()
                    time.sleep(1)
                    inputTag_width.send_keys(str(分辨率A))
                    inputTag_width.click()
                    time.sleep(1)
                    inputTag_height.send_keys(str(分辨率B))

                    # 关闭Excel文件
                    workbook.close()

                    # 点击输入框等待1秒
                    inputTag.click()
                    time.sleep(1)
                    # 将组合好的文本输入到文本框
                    inputTag.send_keys(Keys.CONTROL+"a")
                    inputTag.send_keys(input_text)

                    # 在这输出你的prompt注释，不想输出可以在print前加入# 号不输出，可以少写，例如不写画风，自行选择
                    print("————————————————————————————————————————————————————————————————")
                    print("本次随机结果是：")
                    print(f"{description}")
                    print(f"生成分辨率：{分辨率A} x {分辨率B}")
                    print("————————————————————————————————————————————————————————————————")

                    # 在这修改随机循环时间
                    time.sleep(180)
                except:
                    pass
        except:
            pass

    # 定义点击按钮的函数(不用动)
    def click_button():
        count = 1
        clicks_per_generation = 10
        while True:
            try:
                # 尝试点击按钮
                driver.find_element(By.XPATH, '//*[@id="__next"]/div[2]/div[4]/div[1]/div[5]/button').click()
                # 输出点击次数
                if count % clicks_per_generation == 0:
                    print(f"生成点击次数: {count}")
                # 等待一段时间
                time.sleep(10)
                # 增加计数
                count += 1
                pass
            except:
                pass


    # 下载并刷新网页
    def download():
        count1 = 1
        clicks_per_generation = 10
        while True:
            try:
               # 调整下载并刷新网页的时间，推荐值（tag更换时间的三分之一或者十分之一，再减去80秒。例如1200秒的tag更换时间=1200/3-80=320）
                time.sleep(300)
                # 尝试点击按钮
                while True:
                    # 查找按钮并点击
                    try:
                        driver.find_element(By.XPATH, '//*[@id="historyContainer"]/button').click()
                        WebDriverWait(driver, 10).until(EC.alert_is_present())
                        break
                    except:
                        pass
                alert = Alert(driver)
                alert.accept()
                time.sleep(20)
                while True:
                    try:
                        # 刷新网页
                        driver.refresh()
                        WebDriverWait(driver, 8).until(EC.alert_is_present())
                        break
                        pass
                    except:
                        pass
                # 处理弹出的警告框
                alert = driver.switch_to.alert
                # Accept the alert
                alert.accept()
                time.sleep(60)
                # 输出点击次数
                if count1 % clicks_per_generation == 0:
                    print(f"生成下载次数: {count1}")
                # 增加计数
                count1 += 1
                # 用于处理可能出现的刷新网页后分辨率重制
                inputTag_width = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[3]/div/div[2]/div[2]/input[1]')
                inputTag_height = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[3]/div/div[2]/div[2]/input[2]')
                inputTag_width.click()
                time.sleep(1)
                inputTag_width.send_keys(str(resolution_A))
                inputTag_width.click()
                time.sleep(1)
                inputTag_height.send_keys(str(resolution_B))
                time.sleep(3)

                pass

            except:

                pass

    #检测咒语是否正常,不用动
    def examine():
        global input_text
        while True:
            try:
                inputTag = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[2]/div/div[2]/textarea')
                input_textNOW = inputTag.get_attribute("value")
                if  input_textNOW != input_text:
                    inputTag.click()
                    time.sleep(1)
                    inputTag.send_keys(Keys.CONTROL + "a")
                    inputTag.send_keys(input_text)
                else:
                    time.sleep(20)
                    pass
            except:
                    pass

    #检测网络波动导致掉线的情况
    def examineweb():
        time.sleep(300)
        while True:
            try:
                now_url = driver.current_url
                if  now_url != 'https://novelai.net/image':
                    time.sleep(20)
                    now_now_url = driver.current_url
                    if now_now_url != 'https://novelai/image':
                        python = sys.executable
                        os.execl(python, python, *sys.argv)
                    else:
                        continue
                else:
                    time.sleep(300)
                    pass
            except:
                    pass

    #检测网络波动导致长时间转圈的情况
    def examine_refresh():
        attempt = 0
        max_attempts = 2
        time.sleep(300)
        while True:
            try:
                element = driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/div[4]/div[1]/div[3]/div[2]/div/div[2]/textarea')
                time.sleep(100)
            except:
                time.sleep(20)
                attempt += 1
                if attempts == max_attempts:
                    # 刷新网页
                    print('Refreshing page...')
                    driver.refresh()
                    time.sleep(2)
                    driver.refresh()
                    time.sleep(2)
                    driver.refresh()
                    attempts = 0
                pass

    def ziprename():

        # 压缩包的下载位置，与上方保持一致
        source_folder = 'E:\\nai3自动化\\novelaitemp'
        # 压缩包最终存放的位置
        destination_folder = 'E:\\nai3自动化\\novelai'
        if not os.path.exists('E:\\nai3自动化\\novelai'):
            os.mkdir('E:\\nai3自动化\\novelai')

        while True:
            try:
            # 重命名字符串
                global description
                new_name = f'{description}'
            # 获取当前时间
                current_time = time.strftime('%H-%M-%S')

            # 遍历源文件夹内所有zip文件
                for file_name in os.listdir(source_folder):
                    if file_name.endswith('.zip'):                    # 构造新文件名
                        new_file_name = f'{new_name}_{current_time}.zip'

                    # 重命名文件
                        os.rename(os.path.join(source_folder, file_name), os.path.join(source_folder, new_file_name))

                    # 移动文件
                        shutil.move(os.path.join(source_folder, new_file_name),
                                os.path.join(destination_folder, new_file_name))

                    time.sleep(20)
                    pass
            except:
                    pass

    thread_read_text = threading.Thread(target=read_text)
    thread_read_text.start()
    thread_click_button = threading.Thread(target=click_button)
    thread_click_button.start()
    thread_download = threading.Thread(target=download)
    thread_download.start()
    thread_ziprename = threading.Thread(target=ziprename)
    thread_ziprename.start()
    thread_timer = threading.Thread(target=run_timer)
    thread_timer.start()
    thread_examine= threading.Thread(target=examine)
    thread_examine.start()

    thread_examineweb= threading.Thread(target=examineweb)
    thread_examineweb.start()

    thread_examine_refresh= threading.Thread(target=examine_refresh())
    thread_examine_refresh.start()

    thread_read_text.join()
    thread_examine.join()
    thread_examineweb.join()
    thread_examine_refresh.join()
    thread_timer.join()
    thread_click_button.join()
    thread_download.join()
    thread_ziprename.join()

if __name__ == '__main__':
    run_webdriver()
